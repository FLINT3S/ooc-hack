from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class XlsxWriter:
    _counter: int = 2
    _wb: Workbook
    _worksheet: Worksheet
    _fields_involvement: dict[str, bool]
    _fields_titles = {
        "realtyAddress": "Адрес",
        "realtyType": "Тип",
        "realtyOwner": "Собственник",
        "taskTitle": "Наименование решения",
        "taskDeadline": "Срок исполнения",
        "taskStatus": "Статус",
        "taskAssignee": "Ответственный"
    }
    _value_getters = {
        "realtyAddress": lambda task: task["realEstate"]["address"],
        "realtyType": lambda task: task["realEstate"]["buildingType"]["title"],
        "realtyOwner": lambda task: task["realEstate"]["owner"],
        "taskTitle": lambda task: task["title"],
        "taskDeadline": lambda task: task["deadline"],
        "taskStatus": lambda task: task["status"],
        "taskAssignee": lambda task: f"{task['assignee']['name']} {task['assignee']['surname']}, "
                                     f"{task['assignee']['email']}"
    }

    def __init__(self, fields_involvement):
        self._fields_involvement = fields_involvement
        self._wb = Workbook()
        self._configure_worksheet()
        self._configure_columns()

    def _configure_worksheet(self):
        self._worksheet = self._wb.active
        self._worksheet.title = "Промышленные территории"

    def _configure_columns(self):
        self._worksheet.cell(1, 1, "№ п/п")
        column = 2
        for title in self._fields_titles.keys():
            if self._fields_involvement[title]:
                self._worksheet.cell(1, column, self._fields_titles[title])
                column += 1

    def generate_tasks(self, tasks):
        for task in tasks:
            self._write_task(task)
            self._counter += 1
        self._resize_columns()

    def _write_task(self, task: dict):
        self._worksheet.cell(self._counter, 1, self._counter - 1)
        column = 2
        for getter_name in self._value_getters.keys():
            if self._fields_involvement[getter_name]:
                self._worksheet.cell(self._counter, column, self._value_getters[getter_name](task))
                column += 1

    def _resize_columns(self):
        for col in self._worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.coordinate in self._worksheet.merged_cells:
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            self._worksheet.column_dimensions[column].width = adjusted_width

    def save(self, path: str):
        self._wb.save(path)
